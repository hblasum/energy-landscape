import requests
import openpyxl

# last_activity_at

my_username = 'hblasum'
my_token = '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'

def extract_username_and_repo(repo_url):
    # Extract username and repo name from the URL
    parts = repo_url.split("/")
    if len(parts) < 4 or parts[2] != "github.com":
        return None
    return parts[3], parts[4]

def identify_license(license_text):
  """
  Attempts to identify a common open source license based on keywords.

  Args:
      license_text (str): The text content of the license file.

  Returns:
      str: The possible license name or "Unknown" if not found.
  """
  licenses = {
      "MIT License": ["MIT", "MIT License"],
      "Apache License": ["Apache", "Apache License"],
      "GNU General Public License": ["GPL", "GNU General Public License"],
      "GNU Affero General Public License v3.0": ["Affero", "GNU Affero General Public License v3.0"],
      "BSD License": ["BSD", "BSD License"],
      "Mozilla Public License": ["MPL", "Mozilla Public License"]
  }

  for license, keywords in licenses.items():
    for keyword in keywords:
      if keyword.lower() in license_text.lower():
        return license

  return "Unknown"

def get_github_license(username, repo_name):
  """Fetches the license information of a GitHub project.

  Args:
      username: The username of the GitHub account that owns the repository.
      repo_name: The name of the repository.

  Returns:
      A dictionary containing information about the license, or None if not found.
  """

  url = f"https://api.github.com/repos/{username}/{repo_name}/license"

  response = requests.get(url, auth=(my_username,my_token))

  if response.status_code == 200:
    license_file = response.json()
    # response = requests.get(f"{url}/{license_file}" , auth=(my_username, my_token))
    try:
        return license_file['license']
    except:
        print("Exception in license name lookup")
        return None
  else:
    print(f"Error retrieving license: {response.status_code}")
    return None


# Load the input Excel file
workbook = openpyxl.load_workbook("input.xlsx")
sheet = workbook.active

# Add a new column for licenses
sheet.insert_cols(2)  # Insert a new column at index 2 (after "Repo")
sheet.cell(row=1, column=3).value = "License"  # Set header for the new column

# Fetch license information and update the sheet
for row in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
    repo_url = sheet.cell(row, 1).value

    if repo_url:
        try:
            username, repo_name = extract_username_and_repo(repo_url)
            license_info = get_github_license(username, repo_name)

            if license_info:
                sheet.cell(row, 2).value = license_info['name']
            else:
                sheet.cell(row, 2).value = "Not found"
        except Exception as e:
            print(f"Error processing {repo_url}: {e}")
            sheet.cell(row, 2).value = "Error"
    else:
        print(f"Skipping empty cell in row {row}")

# Save the modified workbook as output.xlsx
workbook.save("output.xlsx")

