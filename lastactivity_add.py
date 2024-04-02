import requests
import openpyxl
from datetime import datetime

# last_activity_at

my_username = 'hblasum'
my_token = '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'

def extract_username_and_repo(repo_url):
    # Extract username and repo name from the URL
    parts = repo_url.split("/")
    if len(parts) < 4 or parts[2] != "github.com":
        return None
    return parts[3], parts[4]


def data_filter (data):
    out = []
    for d in data:
        if datetime.strptime(d["commit"]["committer"]["date"], "%Y-%m-%dT%H:%M:%SZ") >= datetime(2023,1,1):
            out.append(d)
    return out

def get_github_commit_activity(username, repo_name):
  url = f"https://api.github.com/repos/{username}/{repo_name}/commits?per_page=100"

  print(f"Processing: {repo_name}")

  response = requests.get(url, auth=(my_username,my_token))

  if response.status_code == 200:
      page = 1
      data = response.json()
      if data is None:
            return None
      commit_date = data[0]["commit"]["committer"]["date"]
      # Convert date to a more readable format (optional)

      formatted_date = datetime.strptime(commit_date, "%Y-%m-%dT%H:%M:%SZ").strftime("%Y-%m-%d")
      print(f"Last commit date: {formatted_date}")
      data = data_filter(data)
      if data is None:
          return (formatted_date, None, None)

      page_commits = len(data)
      while page_commits == 100 and page < 10:
          page_commits = 0
          page += 1
          print(f"looking at page {page}: ", end="")
          url = f"https://api.github.com/repos/{username}/{repo_name}/commits?per_page=100&page={page}"
          response = requests.get(url, auth=(my_username, my_token))
          if response is not None:
            data_input = response.json()
            print(data_input[0]["commit"]["committer"]["date"])
            new_data = data_filter(response.json())
            page_commits = len(data)
            data += new_data
      total_commits = len(data)
      print(f"Total commits: {total_commits}")
      # print(data[0])
      unique_authors = {}
      for d in data:
        try:
            # print(d["author"]["login"])
            unique_authors[d["author"]["login"]] = 1
        except Exception as e:
            print(f"Error processing author")

      num_authors = len(unique_authors)
      print(f"Number of unique committers: {num_authors}")
      return str(formatted_date), str(total_commits), str(num_authors)
  else:
      print(f"Error retrieving data: {response.status_code}")
      return None


# Load the input Excel file
workbook = openpyxl.load_workbook("input.xlsx")
sheet = workbook.active

# Add new cols
sheet.insert_cols(2)  # Insert a new column at index 2 (after "Repo")
sheet.insert_cols(2)
sheet.insert_cols(2)
sheet.cell(row=1, column=2).value = "Last activity"  # Set header for the new column
sheet.cell(row=1, column=3).value = "Total commits"  # Set header for the new column
sheet.cell(row=1, column=4).value = "Number of authors"  # Set header for the new column


# Fetch license information and update the sheet
for row in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
    repo_url = sheet.cell(row, 1).value

    if repo_url:
        try:
            username, repo_name = extract_username_and_repo(repo_url)
            activity = get_github_commit_activity(username, repo_name)
            if activity is None:
                continue
            (last_activity_info, total_commits_info, num_authors_info) = activity

            if last_activity_info:
                sheet.cell(row, 2).value = last_activity_info
            else:
                sheet.cell(row, 2).value = "Not found"
            if total_commits_info:
                sheet.cell(row, 3).value = total_commits_info
            else:
                sheet.cell(row, 3).value = "Not found"
            if num_authors_info:
                sheet.cell(row, 4).value = num_authors_info
            else:
                sheet.cell(row, 4).value = "Not found"
        except Exception as e:
            print(f"Error processing {repo_url}: {e} lai {last_activity_info} tci {total_commits_info} nai {num_authors_info}")
            sheet.cell(row, 2).value = "Error"
    else:
        print(f"Skipping empty cell in row {row}")

# Save the modified workbook as output.xlsx
workbook.save("output.xlsx")

